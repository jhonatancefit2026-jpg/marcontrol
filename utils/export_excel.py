"""
Exportación a Excel usando openpyxl
"""
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from datetime import datetime
import tkinter.filedialog as filedialog
import tkinter.messagebox as messagebox


# Colores corporativos MarControl
COLOR_HEADER = "1A56DB"
COLOR_SUBHEADER = "E8F0FE"
COLOR_ALT_ROW = "F7FAFC"
COLOR_BORDER = "CBD5E0"
COLOR_TITLE = "0F172A"


def _borde_delgado():
    lado = Side(style='thin', color=COLOR_BORDER)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _borde_grueso():
    lado = Side(style='medium', color=COLOR_HEADER)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def exportar_a_excel(titulo, columnas, datos, filtro_info=None, nombre_sugerido=None):
    """
    Exporta datos a un archivo Excel con formato profesional.
    
    :param titulo: Título del reporte
    :param columnas: Lista de nombres de columnas
    :param datos: Lista de listas/tuplas con los datos
    :param filtro_info: Texto descriptivo del filtro aplicado
    :param nombre_sugerido: Nombre de archivo sugerido
    """
    ruta = filedialog.asksaveasfilename(
        title="Guardar como Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        initialfile=nombre_sugerido or f"{titulo.replace(' ', '_')}.xlsx"
    )
    if not ruta:
        return False

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = titulo[:31]  # Max 31 chars para nombre de hoja

        # ── TÍTULO PRINCIPAL ──────────────────────────────────────
        ws.merge_cells(f"A1:{get_column_letter(len(columnas))}1")
        celda_titulo = ws.cell(row=1, column=1, value=f"MarControl — {titulo}")
        celda_titulo.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        celda_titulo.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        # ── SUBTÍTULO / FECHA ─────────────────────────────────────
        ws.merge_cells(f"A2:{get_column_letter(len(columnas))}2")
        fecha_gen = datetime.now().strftime("%d/%m/%Y %H:%M")
        subtitulo = f"Generado: {fecha_gen}"
        if filtro_info:
            subtitulo += f"  |  Filtro: {filtro_info}"
        celda_sub = ws.cell(row=2, column=1, value=subtitulo)
        celda_sub.font = Font(name="Calibri", size=10, italic=True, color="4A5568")
        celda_sub.fill = PatternFill("solid", fgColor="EBF4FF")
        celda_sub.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20

        # ── CABECERAS DE COLUMNAS ─────────────────────────────────
        for col_idx, col_nombre in enumerate(columnas, start=1):
            celda = ws.cell(row=3, column=col_idx, value=col_nombre.upper())
            celda.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="2D3748")
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.border = _borde_delgado()
        ws.row_dimensions[3].height = 24

        # ── DATOS ─────────────────────────────────────────────────
        for fila_idx, fila in enumerate(datos, start=4):
            color_fondo = "FFFFFF" if (fila_idx % 2 == 0) else COLOR_ALT_ROW
            for col_idx, valor in enumerate(fila, start=1):
                celda = ws.cell(row=fila_idx, column=col_idx, value=str(valor) if valor is not None else "")
                celda.font = Font(name="Calibri", size=10)
                celda.fill = PatternFill("solid", fgColor=color_fondo)
                celda.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                celda.border = _borde_delgado()
            ws.row_dimensions[fila_idx].height = 18

        # ── TOTALES ───────────────────────────────────────────────
        fila_total = len(datos) + 4
        ws.merge_cells(f"A{fila_total}:{get_column_letter(len(columnas))}{ fila_total}")
        celda_tot = ws.cell(row=fila_total, column=1,
                            value=f"Total de registros: {len(datos)}")
        celda_tot.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        celda_tot.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        celda_tot.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[fila_total].height = 18

        # ── ANCHO DE COLUMNAS AUTOMÁTICO ──────────────────────────
        for col_idx, col_nombre in enumerate(columnas, start=1):
            max_ancho = len(col_nombre) + 4
            for fila in datos:
                val = str(fila[col_idx - 1]) if col_idx - 1 < len(fila) else ""
                if len(val) > max_ancho:
                    max_ancho = len(val)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_ancho + 2, 50)

        # ── CONGELAR CABECERA ─────────────────────────────────────
        ws.freeze_panes = "A4"

        wb.save(ruta)
        messagebox.showinfo("Exportación exitosa",
                            f"Archivo guardado:\n{ruta}\n\n{len(datos)} registros exportados.")
        return True

    except Exception as e:
        messagebox.showerror("Error al exportar", f"No se pudo guardar el archivo:\n{e}")
        return False


def exportar_filtrado(titulo, columnas, datos_completos, campo_fecha_idx=None,
                      campo_categoria_idx=None, nombre_sugerido=None):
    """
    Exporta con filtros de fecha y/o categoría.
    Abre un diálogo de filtros antes de exportar.
    """
    from tkinter import Toplevel, Label, Entry, Button, StringVar, Frame, ttk
    import tkinter as tk

    ventana = Toplevel()
    ventana.title("Filtros de exportación")
    ventana.geometry("400x300")
    ventana.grab_set()
    ventana.resizable(False, False)

    resultado = {'datos': datos_completos, 'filtro': None}

    Label(ventana, text="Filtros de exportación", font=("Calibri", 14, "bold")).pack(pady=10)

    frame = Frame(ventana, padx=20)
    frame.pack(fill="x")

    var_desde = StringVar()
    var_hasta = StringVar()
    var_cat = StringVar()

    if campo_fecha_idx is not None:
        Label(frame, text="Fecha desde (DD/MM/YYYY):").grid(row=0, column=0, sticky="w", pady=4)
        Entry(frame, textvariable=var_desde, width=20).grid(row=0, column=1, padx=5)
        Label(frame, text="Fecha hasta (DD/MM/YYYY):").grid(row=1, column=0, sticky="w", pady=4)
        Entry(frame, textvariable=var_hasta, width=20).grid(row=1, column=1, padx=5)

    if campo_categoria_idx is not None:
        categorias = sorted(set(str(f[campo_categoria_idx]) for f in datos_completos if f[campo_categoria_idx]))
        Label(frame, text="Categoría:").grid(row=2, column=0, sticky="w", pady=4)
        combo = ttk.Combobox(frame, textvariable=var_cat, values=["Todas"] + categorias, width=18)
        combo.set("Todas")
        combo.grid(row=2, column=1, padx=5)

    def aplicar():
        datos_filtrados = list(datos_completos)
        info_filtro = []

        if campo_fecha_idx is not None:
            from utils.validators import parsear_fecha
            desde = var_desde.get().strip()
            hasta = var_hasta.get().strip()
            if desde:
                d_desde = parsear_fecha(desde)
                if d_desde:
                    datos_filtrados = [f for f in datos_filtrados
                                       if parsear_fecha(str(f[campo_fecha_idx])) and
                                       parsear_fecha(str(f[campo_fecha_idx])) >= d_desde]
                    info_filtro.append(f"Desde {desde}")
            if hasta:
                d_hasta = parsear_fecha(hasta)
                if d_hasta:
                    datos_filtrados = [f for f in datos_filtrados
                                       if parsear_fecha(str(f[campo_fecha_idx])) and
                                       parsear_fecha(str(f[campo_fecha_idx])) <= d_hasta]
                    info_filtro.append(f"Hasta {hasta}")

        if campo_categoria_idx is not None:
            cat = var_cat.get()
            if cat and cat != "Todas":
                datos_filtrados = [f for f in datos_filtrados
                                   if str(f[campo_categoria_idx]) == cat]
                info_filtro.append(f"Categoría: {cat}")

        resultado['datos'] = datos_filtrados
        resultado['filtro'] = " | ".join(info_filtro) if info_filtro else None
        ventana.destroy()

    def sin_filtro():
        ventana.destroy()

    frame_btn = Frame(ventana)
    frame_btn.pack(pady=20)
    Button(frame_btn, text="Aplicar filtro", command=aplicar,
           bg="#1A56DB", fg="white", width=14).pack(side="left", padx=5)
    Button(frame_btn, text="Sin filtro", command=sin_filtro,
           bg="#4A5568", fg="white", width=14).pack(side="left", padx=5)

    ventana.wait_window()

    exportar_a_excel(titulo, columnas, resultado['datos'],
                     filtro_info=resultado['filtro'],
                     nombre_sugerido=nombre_sugerido)
